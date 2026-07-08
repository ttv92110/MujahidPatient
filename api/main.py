from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import FileResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pathlib import Path
from datetime import datetime
import json
import shutil
import os
import io
import pickle
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====== GOOGLE DRIVE OAuth IMPORTS ======
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# ====== BASE PATHS ======
BASE_DIR = Path(__file__).parent.parent

if os.environ.get("VERCEL"):
    LOCAL_DATA_DIR = Path("/tmp/data")
else:
    LOCAL_DATA_DIR = BASE_DIR / "data"

TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"
CLIENT_SECRETS_FILE = BASE_DIR / "api" / "client_secrets.json"
TOKEN_FILE = BASE_DIR / "token.pickle"

LOCAL_DATA_DIR.mkdir(exist_ok=True)
TEMPLATES_DIR.mkdir(exist_ok=True)
STATIC_DIR.mkdir(exist_ok=True)

# ====== APP ======
app = FastAPI(title="Patient Data Entry")

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# ============================================================
# GOOGLE DRIVE AUTH (OAuth 2.0)
# ============================================================
SCOPES = ["https://www.googleapis.com/auth/drive.file"]  # درست کیا گیا

DRIVE_FOLDER_NAME = "Patient Data"  # فولڈر کا نام

def get_drive_service():
    creds = None
    if TOKEN_FILE.exists():
        with open(TOKEN_FILE, "rb") as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(GoogleAuthRequest())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CLIENT_SECRETS_FILE), SCOPES)
            creds = flow.run_local_server(port=0, open_browser=True)
        with open(TOKEN_FILE, "wb") as token:
            pickle.dump(creds, token)
    return build("drive", "v3", credentials=creds)

# ====== FIND OR CREATE FOLDER ======
_folder_id_cache = None  # پہلی بار حاصل کرنے کے بعد کیش میں رکھیں

def get_or_create_drive_folder(service):
    """Drive پر 'Patient Data' فولڈر ڈھونڈیں یا بنائیں اور اس کی ID واپس کریں"""
    global _folder_id_cache
    if _folder_id_cache:
        return _folder_id_cache

    # تلاش کریں کہ کیا فولڈر پہلے سے موجود ہے
    query = f"name='{DRIVE_FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    folders = results.get("files", [])

    if folders:
        folder_id = folders[0]["id"]
    else:
        # فولڈر موجود نہیں، نیا بنائیں
        file_metadata = {
            "name": DRIVE_FOLDER_NAME,
            "mimeType": "application/vnd.google-apps.folder"
        }
        folder = service.files().create(body=file_metadata, fields="id").execute()
        folder_id = folder["id"]

    _folder_id_cache = folder_id
    return folder_id

# ============================================================
# DRIVE SYNC FUNCTIONS (OAuth) - Folder Aware
# ============================================================
def find_or_create_drive_file(service, month_year_str, folder_id):
    filename = f"Patient List {month_year_str}.xlsx"
    # مخصوص فولڈر میں تلاش کریں
    query = f"name='{filename}' and '{folder_id}' in parents and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    else:
        file_metadata = {
            "name": filename,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "parents": [folder_id]  # فولڈر میں محفوظ کریں
        }
        file = service.files().create(body=file_metadata, fields="id").execute()
        return file["id"]

def upload_to_drive(service, file_id, filepath):
    try:
        media = MediaIoBaseUpload(
            io.BytesIO(filepath.read_bytes()),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True
        )
        service.files().update(fileId=file_id, media_body=media).execute()
        return True
    except Exception as e:
        print(f"❌ Upload failed: {e}")
        return False

def download_from_drive(service, file_id, filepath):
    try:
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        filepath.write_bytes(fh.getvalue())
        return True
    except Exception as e:
        print(f"❌ Download failed: {e}")
        return False

def sync_drive_to_local(month_year_str):
    service = get_drive_service()
    if not service:
        return False
    folder_id = get_or_create_drive_folder(service)
    filepath = LOCAL_DATA_DIR / f"Patient List {month_year_str}.xlsx"
    file_id = find_or_create_drive_file(service, month_year_str, folder_id)
    if file_id:
        return download_from_drive(service, file_id, filepath)
    return False

def sync_local_to_drive(month_year_str):
    service = get_drive_service()
    if not service:
        return False
    filepath = LOCAL_DATA_DIR / f"Patient List {month_year_str}.xlsx"
    if not filepath.exists():
        return False
    folder_id = get_or_create_drive_folder(service)
    file_id = find_or_create_drive_file(service, month_year_str, folder_id)
    if file_id:
        return upload_to_drive(service, file_id, filepath)
    return False

# ============================================================
# EXCEL STYLING & FILE OPERATIONS
# ============================================================
def style_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_length + 4, 35)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    wb.save(filepath)

def create_new_month_file(month_year_str):
    filepath = LOCAL_DATA_DIR / f"Patient List {month_year_str}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:S4')
    title = f"Patients List ({get_month_year_with_hyphen(month_year_str)})"
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name='Calibri', size=18, bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    headers = [
        "Sr. No", "DATE", "NAME", "AGE", "CATEGORY", "MR#NO",
        "CONSULTANT NAME", "PROCEDURE", "TOTAL COST", "11% TAX Deduction",
        "Total Cost After 11% TAX", "PAYMENT", "RECEIVED", "DISPOSABLE",
        "CATH SHARE", "DOC SHARE", "SSP CLAIM", "Status",
        "Scan Done/ Upload in State Life Web Portal"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=14, bold=True, color="000000")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 30
    wb.save(filepath)
    style_excel(filepath)
    return filepath

def get_filepath(month_year_str):
    return LOCAL_DATA_DIR / f"Patient List {month_year_str}.xlsx"

def append_row_to_excel(month_year_str, row_data):
    filepath = get_filepath(month_year_str)
    if not filepath.exists():
        create_new_month_file(month_year_str)
    wb = load_workbook(filepath)
    ws = wb.active
    next_row = ws.max_row + 1
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col_idx, value=value)
    wb.save(filepath)
    style_excel(filepath)
    sync_local_to_drive(month_year_str)

def update_row_in_excel(month_year_str, sr_no, row_data):
    filepath = get_filepath(month_year_str)
    if not filepath.exists():
        return False
    wb = load_workbook(filepath)
    ws = wb.active
    target_row = None
    for row_idx in range(6, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == sr_no:
            target_row = row_idx
            break
    if target_row is None:
        return False
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=target_row, column=col_idx, value=value)
    wb.save(filepath)
    style_excel(filepath)
    sync_local_to_drive(month_year_str)
    return True

def delete_row_from_excel(month_year_str, sr_no):
    filepath = get_filepath(month_year_str)
    if not filepath.exists():
        return False
    wb = load_workbook(filepath)
    ws = wb.active
    target_row = None
    for row_idx in range(6, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == sr_no:
            target_row = row_idx
            break
    if target_row is None:
        return False
    ws.delete_rows(target_row)
    wb.save(filepath)
    style_excel(filepath)
    sync_local_to_drive(month_year_str)
    return True

def get_all_rows(month_year_str):
    filepath = get_filepath(month_year_str)
    if not filepath.exists():
        sync_drive_to_local(month_year_str)
        if not filepath.exists():
            return []
    wb = load_workbook(filepath)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row and any(row):
            rows.append(list(row))
    return rows


# ============================================================
# HELPER: Capitalize Words
# ============================================================
def capitalize_words(text: str) -> str:
    if not text:
        return text
    text = text.strip()
    if text.lower().startswith("dr."):
        rest = text[3:].strip()
        if rest:
            rest = ' '.join(word.capitalize() for word in rest.split())
            return f"Dr. {rest}"
        else:
            return "Dr."
    else:
        return ' '.join(word.capitalize() for word in text.split())

# ============================================================
# SUGGESTIONS
# ============================================================
def build_suggestions():
    names = set()
    consultants = {"Dr. Anjum Rana"}
    disposables = set()
    for filepath in LOCAL_DATA_DIR.glob("Patient List *.xlsx"):
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=6, values_only=True):
                if row and any(row):
                    name = row[2] if len(row) > 2 else None
                    consultant = row[6] if len(row) > 6 else None
                    disposable = row[13] if len(row) > 13 else None
                    if name:
                        names.add(capitalize_words(str(name)))
                    if consultant:
                        consultants.add(capitalize_words(str(consultant)))
                    if disposable:
                        disposables.add(capitalize_words(str(disposable)))
        except Exception as e:
            print(f"Error reading {filepath}: {e}")
    return {
        "names": sorted(list(names)),
        "consultants": sorted(list(consultants)),
        "disposables": sorted(list(disposables))
    }

suggestions_cache = build_suggestions()

def get_suggestions():
    return suggestions_cache

# ============================================================
# OTHER HELPER FUNCTIONS
# ============================================================
def get_month_year_from_date(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return f"{dt.strftime('%B')} {dt.year}"

def get_month_year_with_hyphen(month_year_str):
    parts = month_year_str.split()
    if len(parts) == 2:
        return f"{parts[0]}- {parts[1]}"
    return month_year_str

def parse_date_dmy(date_str):
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except Exception:
        return datetime.min

def format_date_dmy(date_str):
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return date_str

def get_last_sr_no(month_year_str):
    filepath = get_filepath(month_year_str)
    if filepath.exists():
        wb = load_workbook(filepath)
        ws = wb.active
        max_row = ws.max_row
        if max_row >= 6:
            last_row = ws[max_row]
            sr_no = last_row[0].value
            if sr_no and isinstance(sr_no, int):
                return sr_no
        return 0
    return 0

# ============================================================
# ENDPOINTS
# ============================================================

@app.get("/")
async def home(request: Request, date: Optional[str] = Query(None)):
    suggestions = get_suggestions()
    if date:
        selected_date = date
    else:
        selected_date = datetime.now().strftime("%Y-%m-%d")
    return templates.TemplateResponse("index.html", {
        "request": request,
        "names": suggestions["names"],
        "consultants": suggestions["consultants"],
        "disposables": suggestions["disposables"],
        "default_consultant": "Dr. Anjum Rana",
        "selected_date": selected_date
    })


@app.post("/submit")
async def submit(
    request: Request,
    date: str = Form(...),
    name: str = Form(...),
    age: str = Form(...),
    category: str = Form(...),
    mr_no: str = Form(""),
    consultant: str = Form(...),
    procedure: str = Form(...),
    total_cost: float = Form(...),
    tax: float = Form(...),
    total_after: float = Form(...),
    payment: str = Form(""),
    received: str = Form(""),
    disposable: str = Form(...),
    cath_share: str = Form(""),
    doc_share: str = Form(""),
    ssp_claim: str = Form(""),
    status: str = Form(""),
    scan_done: str = Form(""),
    sr_no_hidden: Optional[int] = Form(None),
):
    name = capitalize_words(name)
    consultant = capitalize_words(consultant)
    disposable = capitalize_words(disposable)
    
    month_year = get_month_year_from_date(date)
    date_dmy = format_date_dmy(date)
    
    if sr_no_hidden:
        row_data = [
            sr_no_hidden, date_dmy, name, age, category, mr_no,
            consultant, procedure, total_cost, tax,
            total_after, payment, received, disposable,
            cath_share, doc_share, ssp_claim, status, scan_done
        ]
        update_row_in_excel(month_year, sr_no_hidden, row_data)
    else:
        last_sr = get_last_sr_no(month_year)
        new_sr = last_sr + 1
        row_data = [
            new_sr, date_dmy, name, age, category, mr_no,
            consultant, procedure, total_cost, tax,
            total_after, payment, received, disposable,
            cath_share, doc_share, ssp_claim, status, scan_done
        ]
        append_row_to_excel(month_year, row_data)
    
    global suggestions_cache
    suggestions_cache = build_suggestions()
    
    return RedirectResponse(f"/?date={date}", status_code=303)

@app.get("/files")
async def list_files():
    files = []
    for filepath in LOCAL_DATA_DIR.glob("Patient List *.xlsx"):
        name = filepath.stem
        month_year = name.replace("Patient List ", "")
        files.append({"name": filepath.name, "month_year": month_year})
    files.sort(key=lambda x: x["month_year"], reverse=True)
    return files

@app.get("/preview")
async def preview_data(month_year: str):
    rows = get_all_rows(month_year)
    return rows

@app.post("/delete")
async def delete_row(month_year: str = Form(...), sr_no: int = Form(...)):
    success = delete_row_from_excel(month_year, sr_no)
    global suggestions_cache
    suggestions_cache = build_suggestions()
    return {"success": success}

@app.get("/download")
async def download_file(month_year: str):
    filepath = get_filepath(month_year)
    if not filepath.exists():
        sync_drive_to_local(month_year)
        if not filepath.exists():
            return JSONResponse({"error": f"File {filepath.name} not found."}, status_code=404)
    return FileResponse(path=filepath, filename=filepath.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/arrange")
async def arrange_file(month_year: str):
    try:
        filepath = get_filepath(month_year)
        if not filepath.exists():
            sync_drive_to_local(month_year)
            if not filepath.exists():
                return JSONResponse({"error": f"File {filepath.name} not found."}, status_code=404)
        
        wb = load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[5]]
        data_rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row and any(row):
                data_rows.append(list(row))
        
        if not data_rows:
            return JSONResponse({"error": "No data rows to arrange."}, status_code=400)
        
        data_rows.sort(key=lambda x: (parse_date_dmy(x[1]) if x[1] else datetime.min, x[7] if x[7] else ""))
        for idx, row in enumerate(data_rows, start=1):
            row[0] = idx
        
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.merge_cells('A1:S4')
        title_text = f"Patients List ({get_month_year_with_hyphen(month_year)})"
        title_cell = new_ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(name='Calibri', size=18, bold=True, color="000000")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, h in enumerate(headers, 1):
            cell = new_ws.cell(row=5, column=col_idx, value=h)
            cell.font = Font(name='Calibri', size=14, bold=True, color="000000")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        new_ws.row_dimensions[5].height = 30
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                new_ws.cell(row=row_idx, column=col_idx, value=value)
        
        backup_path = filepath.with_suffix(".backup.xlsx")
        if filepath.exists():
            shutil.move(filepath, backup_path)
        new_wb.save(filepath)
        style_excel(filepath)
        sync_local_to_drive(month_year)
        return {"message": f"File arranged successfully. Old file backed up as {backup_path.name}."}
    except Exception as e:
        return JSONResponse({"error": f"Arrangement failed: {str(e)}"}, status_code=500)

@app.delete("/delete-file")
async def delete_file(month_year: str):
    filepath = get_filepath(month_year)
    if not filepath.exists():
        return JSONResponse({"error": "File not found"}, status_code=404)
    filepath.unlink()
    backup = filepath.with_suffix(".backup.xlsx")
    if backup.exists():
        backup.unlink()
    service = get_drive_service()
    if service:
        try:
            folder_id = get_or_create_drive_folder(service)
            file_id = find_or_create_drive_file(service, month_year, folder_id)
            service.files().delete(fileId=file_id).execute()
        except Exception as e:
            print(f"❌ Failed to delete from Drive: {e}")
    return {"message": f"File {filepath.name} deleted successfully."}

@app.get("/health")
async def health():
    return {"status": "ok"}

# ============================================================
# ON STARTUP: SYNC DRIVE
# ============================================================
@app.on_event("startup")
async def startup_event():
    print("🔄 Syncing files from Google Drive (OAuth)...")
    try:
        service = get_drive_service()
        if not service:
            print("⚠️ Google Drive not available. Using local storage only.")
            return
        # پہلے فولڈر کو یقینی بنائیں
        folder_id = get_or_create_drive_folder(service)
        # اب فولڈر کے اندر تمام فائلیں ڈھونڈیں
        query = f"'{folder_id}' in parents and trashed=false"
        results = service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get("files", [])
        for file in files:
            if file["name"].startswith("Patient List ") and file["name"].endswith(".xlsx"):
                month_year = file["name"].replace("Patient List ", "").replace(".xlsx", "")
                local_path = LOCAL_DATA_DIR / file["name"]
                if not local_path.exists():
                    download_from_drive(service, file["id"], local_path)
                    print(f"✅ Downloaded {file['name']} from Drive")
        print("✅ Sync complete.")
    except Exception as e:
        print(f"⚠️ Drive sync failed: {e}")
        print("⚠️ Continuing with local storage only.")
        